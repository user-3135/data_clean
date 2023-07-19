import streamlit as st
import pandas as pd
from io import BytesIO
import excel_file
import excel_file_2
import excel_file_3
import excel_file_4 as writer
import excel_6 as new_writer
import comb
import xlsxwriter as xl
import openpyxl
data = 1
st.header('import file')
uploaded_file_2 = st.file_uploader("Upload a Single File")
# data read
try:
    tb_df_1 = pd.read_excel(uploaded_file_2, sheet_name = 'TB') #BS
except:
    st.write('Trial Balance Error')
try:
    cash_flow_df = pd.read_excel(uploaded_file_2, sheet_name = 'CashFlow') #BS
except:
    st.write('Cash Flow Error')
try:
    bs_df = pd.read_excel(uploaded_file_2, sheet_name = 'BS') #BS
except:
    st.write('Balance Sheet Error - Verify the sheet name is BS')
try:
    is_df = pd.read_excel(uploaded_file_2, sheet_name = 'IS') #BS
except:
    st.write('Income Statement Error - Verify the sheet name is IS')
try:
    actual_budget = pd.read_excel(uploaded_file_2, sheet_name = 'Actual-Budget')
except:
    st.write('Verify that the Actual-Budget sheet is called Actual-Budget with no spaces')
try:
    data_ar_detail = pd.read_excel(uploaded_file_2, sheet_name = 'AR Detail')
except:
    st.write('AR Detail Error - Ensure the sheet name is AR Detail')
try:
    data_12_month = pd.read_excel(uploaded_file_2, sheet_name = 'IS 12 Month Actual')
except:
    try:
        data_12_month = pd.read_excel(uploaded_file_2, sheet_name = 'IS 12 Month')
    except:
        try:
            data_12_month = pd.read_excel(uploaded_file_2, sheet_name = ' IS 12 Month')
        except:
            st.write('IS 12 Month Actual - Ensure the sheet name is IS 12 Month Actual')
try:    
    data_ten_sched = pd.read_excel(uploaded_file_2, sheet_name = 'TenSched1')
except:
    try:
        data_ten_sched = pd.read_excel(uploaded_file_2, sheet_name = 'TenSched 1')
    except:
        try:
            data_ten_sched = pd.read_excel(uploaded_file_2, sheet_name = 'TenSche1')
        except:
            st.write('Ten Sched Error - Ensure the sheet name is TenSched1')
# ------------------------------------
try:
    data_payment_register = pd.read_excel(uploaded_file_2, sheet_name = 'Pymnt Register')
except:
    try:
        data_payment_register = pd.read_excel(uploaded_file_2, sheet_name = 'Payment Register') # Payment Register
    except:
        st.write('Pymnt Register Error')
try:
    data_ap_detail = pd.read_excel(uploaded_file_2, sheet_name = 'AP Detail')
except:
    st.write('AP Detail Error')
try:
    data_mth_gl = pd.read_excel(uploaded_file_2, sheet_name = 'MTH GL')
except:
    try:
        data_mth_gl = pd.read_excel(uploaded_file_2, sheet_name = 'MTHGL')
    except:
        try:
            data_mth_gl = pd.read_excel(uploaded_file_2, sheet_name = ' GL')
        except:
            st.write('MTH GL Error - Ensure the sheet name is MTH GL')
try:    
    je_register_data = pd.read_excel(uploaded_file_2, sheet_name = 'JE Register')
except:
    st.write('JE Register Error')
output_4 = BytesIO()
try:
    workbook_not_func = xl.Workbook(output_4)
    #----------------------------------------------------------------------------------- 1
    act_bud_worksheet_func = workbook_not_func.add_worksheet('Actual-Budget')
    Income_Statement_wb = workbook_not_func.add_worksheet('Income Statement')
    worksheet_12_mo_actual = workbook_not_func.add_worksheet('IS 12 Month Actual')
    bs_wb = workbook_not_func.add_worksheet('BS')
    cf_worksheet_func = workbook_not_func.add_worksheet('Cash Flow')
    tb_worksheet_func = workbook_not_func.add_worksheet('Trial Balance')
    worksheet_tenancy_sched = workbook_not_func.add_worksheet('TenSched1')
    aging_detail_sheet = workbook_not_func.add_worksheet('AR Detail')
    worksheet_pay_reg = workbook_not_func.add_worksheet('Payment Register')
    ap_detail_sheet = workbook_not_func.add_worksheet('AP Detail')
    je_register_worksheet = workbook_not_func.add_worksheet('JE Register')
    mnth_gl_worksheet = workbook_not_func.add_worksheet('Mnth GL')
    #----------------------------------------------------------------------------------- 2
    # procs
    try:
        h = new_writer.JE_REGISTER_SHEET(workbook_not_func, je_register_data, je_register_worksheet)
    except:
        try:
            h = writer.JE_REGISTER_SHEET_2(workbook_not_func, je_register_data, je_register_worksheet)
            st.write('used old JE Register Sheet')
        except:
            st.write('error writing JE Register in Excel')
    try:
        f = new_writer.mnth_gl_sheet(workbook_not_func, data_mth_gl, mnth_gl_worksheet) ##mnth_gl_sheet
    except:
        try:
            f = writer.mnth_gl_sheet_2(workbook_not_func, data_mth_gl, mnth_gl_worksheet)
            st.write('used old Mnth GL Function')
        except:
            st.write('error writing Month General Ledger in Excel')
    try:
        e = new_writer.ap_detail_sheet_def_2(workbook_not_func, data_ap_detail, ap_detail_sheet)
    except:
        try:
            e = writer.ap_detail_sheet_def_2(workbook_not_func, data_ap_detail, ap_detail_sheet)
            st.write('used an old AP Detail Sheet')
        except:
            try:
                e = writer.ap_detail_sheet_def(workbook_not_func, data_ap_detail, ap_detail_sheet)
            except:
                st.write('error writing Month AP Detail in Excel')
    try: #payment_register_sheet_2
        d = new_writer.payment_register_sheet_2(workbook_not_func, data_payment_register, worksheet_pay_reg)
    except:
        try:
            d = writer.payment_register_sheet_2(workbook_not_func, data_payment_register, worksheet_pay_reg)
            st.write('used an old payment register sheet')
        except:
            try:
                d = writer.payment_register_sheet(workbook_not_func, data_payment_register, worksheet_pay_reg)
            except:
                st.write('error writing Month Payment Register in Excel')
#----------------------------------------------------------------------------------- 2
    if 1 == 1:
        # procs
        try:
            a1 = new_writer.twelve_month_actual_budget_v2(workbook_not_func, data_12_month, worksheet_12_mo_actual)
        except:
            try:
                a1 = writer.twelve_month_actual_budget(workbook_not_func, data_12_month, worksheet_12_mo_actual)
                st.write('used an old version of the 12 month actual budget')
            except:
                st.write('error 12 month actual budget')
        #-----------------------------------------------------------
        try:
            b1 = new_writer.ten_sched_1_v2(workbook_not_func, data_ten_sched, worksheet_tenancy_sched)
        except:
            try:
                b1 = writer.ten_sched_1(workbook_not_func, data_ten_sched, worksheet_tenancy_sched)
                st.write('used an old version of the ten sched')
            except:
                st.write('ten sched error')
        try:
            c1 = new_writer.aging_detail_2(workbook_not_func, data_ar_detail, aging_detail_sheet)
        except:
            try: #aging_detail_2
                c1 = writer.aging_detail_2(workbook_not_func, data_ar_detail, aging_detail_sheet)
            except:
                try:
                    c1 = writer.aging_detail(workbook_not_func, data_ar_detail, aging_detail_sheet)
                except:
                    st.write('error writing Aging Detail in Excel')
    #----------------------------------------------------------------------------------- 3
    try:
        aa = new_writer.create_xl_cf_v2(workbook_not_func, cash_flow_df, cf_worksheet_func)
    except:
        try:
            aa = writer.create_xl_cf(workbook_not_func, cash_flow_df, cf_worksheet_func)
            st.write('used an old version of the Cash Flow DF')
        except:
            st.write('error writing the Cash Flow Sheet')
    try:
        ba = new_writer.create_xl_tb_v2(workbook_not_func, tb_df_1, tb_worksheet_func)
    except:
        try:
            ba = writer.create_xl_tb(workbook_not_func, tb_df_1, tb_worksheet_func)
            st.write('used an old trial balance sheet')
        except:
            st.write('error writing the trial balance sheet')
    try:
        ca = new_writer.create_xl_balance_sheet_v2(workbook_not_func, bs_df, bs_wb)
    except:
        try:
            ca = writer.create_xl_balance_sheet(workbook_not_func, bs_df, bs_wb)
            st.write('used an old Balance Sheet file')
        except:
            st.write('error writing the balance sheet')
    try:
        fa = new_writer.budget_comp_sheet_creation_v2(workbook_not_func, actual_budget, act_bud_worksheet_func)
    except:
        try:
            fa = writer.budget_comp_sheet_creation(workbook_not_func, actual_budget, act_bud_worksheet_func)
            st.write('used an old budget comp file')
        except:
            st.write('error writing the budget comp sheet')
    try:
        new_writer.income_statement_v2(workbook_not_func, is_df, Income_Statement_wb)
    except:
        try:
            writer.income_statement(workbook_not_func, is_df, Income_Statement_wb)
            st.write('used an old file to write the income statement')
        except:
            st.write('error writing the income statement sheet')
    workbook_not_func.close()
except: 
    st.write('error in base')
try:
    wb_base = openpyxl.load_workbook(output_4)
    vals_to_rip = ['Summary'
                   , 'FS Checklist'
                   , 'Financial Summary Chart'
                   , 'Rent Roll'
                   , 'Mngt Fee '
                   , 'RET & Insurance'
                   , 'Depr Schedule'
                   , 'Amort Schedule'
                   , 'Mortgage Statement'
                   , 'Bank Recon'
                  ]
    wb_add = openpyxl.load_workbook(uploaded_file_2)
    final_output = BytesIO()
    #for i in vals_to_rip:    
    #    target_sheet = wb_base.create_sheet(i)
    #    try:
    #        source_sheet = wb_add[i]
    #        comb.copy_sheet(source_sheet, target_sheet)
    #    except:
    #        print(i)
        
    wb_base.save(final_output)
except:
    pass
# -------------------------------------
try:
    try:
        name = 'Combined Property Workbook Test.xlsx'
        st.download_button(
                            label="Download Excel Combined Workbook",
                            data=final_output.getvalue(),
                            file_name=name,
                            mime="application/vnd.ms-excel"
        )
    except:
        try:
            st.download_button(
                                label="Download Excel Workbook",
                                data=output_4.getvalue(),
                                file_name=name,
                                mime="application/vnd.ms-excel"
            )
        except:
            output_2 = BytesIO()
            excel_file_3.create_excel_v3(actual_budget, output_2, is_df, bs_df, cash_flow_df, tb_df_1 , data_ar_detail, data_12_month, data_ten_sched
                                        , je_register_data, data_mth_gl,data_ap_detail, data_payment_register)
            name = 'Property Workbook Test.xlsx'
            st.download_button(
                                label="Download Excel workbook",
                                data=output_2.getvalue(),
                                file_name=name,
                                mime="application/vnd.ms-excel"
            )
except:
        pass
if data == 2:
        st.header('Upload Budget Comp')
        uploaded_file = st.file_uploader("Choose a file")
        # income_statement_1
        st.header('Upload Income Statement')
        uploaded_file_income_statemnt_1 = st.file_uploader("Upload the Income Statement")
        # bal_sheet_1
        st.header('Upload Balance Sheet Period Change')
        uploaded_file_balance_sheet = st.file_uploader("Upload the Balance Sheet Period Change")
        # cash_flow_1_df
        st.header('Upload Statement of Cashflow')
        uploaded_file_cash_flow = st.file_uploader("Upload the Statement of Cash Flow")
        # trail_balance_df
        st.header('Upload Trial Balance')
        uploaded_file_trial_balance = st.file_uploader("Upload the Trial Balance")
        # payment_register_df
        st.header('Upload Payment Register')
        uploaded_file_payment_register = st.file_uploader("Upload teh Payment Register")
        dataframe_budget_comp = pd.read_excel(uploaded_file)
        dataframe_is = pd.read_excel(uploaded_file_income_statemnt_1)
        dataframe_bs = pd.read_excel(uploaded_file_balance_sheet)
        dataframe_cf = pd.read_excel(uploaded_file_cash_flow)
        dataframe_tb = pd.read_excel(uploaded_file_trial_balance)
        dataframe_pr = pd.read_excel(uploaded_file_payment_register)
        try:
                output = BytesIO()
                excel_file.create_excel_2(dataframe_budget_comp, output, dataframe_is, dataframe_bs, dataframe_cf, dataframe_tb, dataframe_pr)
                name = 'Property Workbook Test.xlsx'
                st.download_button(
                        label="Download Excel workbook",
                        data=output.getvalue(),
                        file_name=name,
                        mime="application/vnd.ms-excel"
                )
        except:
                pass
